import os
import sys
import time
import openpyxl
import pyautogui
from openpyxl.utils import get_column_letter

print("=" * 60)
print("          70-CYCLE TARGETED COLUMN CONTROLLER (V3.1)")
print("=" * 60)

# Local configuration save filename
CONFIG_FILE = "matrix_config.txt"
excel_file = ""

# 1. Attempt to read saved history path from local configuration file
if os.path.exists(CONFIG_FILE):
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            saved_path = f.read().strip()
            if os.path.exists(saved_path):
                print(f"📂 Found saved Excel path:\n   {saved_path}")
                use_saved = input("Use this saved path? (yes/no) [Default: yes]: ").strip().lower()
                if use_saved in ['', 'yes', 'y']:
                    excel_file = saved_path
    except Exception:
        pass  # If reading fails, degrade gracefully to manual entry prompt

# Prompt for manual input if no historical path exists or user rejects the saved path
if not excel_file:
    excel_file = input("\nEnter the full path to your Excel file (e.g., C:\\path\\to\\data.xlsx):\n> ").strip().strip('"\'')

# Validate if the path exists on disk
if not os.path.exists(excel_file):
    print(f"\n❌ ERROR: The specified Excel file could not be found at:\n   {excel_file}")
    sys.exit(1)

# Save the valid path to the local configuration file for subsequent sessions
try:
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        f.write(excel_file)
    print("💾 Path configuration saved locally for next time.")
except Exception as e:
    print(f"⚠️ Warning: Could not save configuration file: {e}")

try:
    # 2. Open spreadsheet using live workbook cell tracking
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # --- Display available sheets and request choice ---
    print("\n📋 Available sheets in this workbook:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {idx}. {sheet_name}")
    print("-" * 60)
    
    selected_sheet = input("Enter the exact name of the sheet you are working on:\n> ").strip()
    
    if selected_sheet not in wb.sheetnames:
        print(f"\n❌ ERROR: '{selected_sheet}' is not a valid sheet name in this workbook.")
        sys.exit(1)
        
    sheet = wb[selected_sheet]

    # --- TARGETED COLUMN SELECTION ROUTINE ---
    print("\n🎯 COLUMN SELECTION TARGETING:")
    print(" Which columns do you want to update in the web browser?")
    print(" Leave blank and press ENTER to update ALL columns (1 to 6).")
    print(" Otherwise, type specific column numbers separated by commas (e.g., 1, 3, 5)")
    col_input = input("Columns to update (1-6):\n> ").strip()

    # Parse targeted column set
    target_columns = set()
    if col_input == "":
        target_columns = {1, 2, 3, 4, 5, 6}  # Default: update everything
        print("✅ Will update ALL 6 columns.")
    else:
        try:
            for part in col_input.split(','):
                num = int(part.strip())
                if 1 <= num <= 6:
                    target_columns.add(num)
                else:
                    raise ValueError
            if not target_columns:
                raise ValueError
            sorted_cols = sorted(list(target_columns))
            print(f"✅ Target Locked. Will ONLY update columns: {sorted_cols}")
            print("ℹ️ Note: Unselected columns will only execute TAB navigation actions.")
        except ValueError:
            print("\n❌ ERROR: Invalid column selection. Please enter numbers between 1 and 6 separated by commas.")
            sys.exit(1)

    # 3. Dynamic row selection prompt (Max extended to 70)
    try:
        start_row = int(input("\nWhich Excel row would you like to start copying from? (1-70):\n> ").strip())
        if start_row < 1 or start_row > 70:
            raise ValueError
    except ValueError:
        print("\n❌ ERROR: Please enter a valid whole number between 1 and 70.")
        sys.exit(1)
        
    # 4. Dynamic cycle requirement prompt (Max extended to 70 bounds)
    max_available_cycles = 70 - start_row + 1
    try:
        requested_cycles = int(input(f"\nHow many rows/cycles do you require to run? (Max available: {max_available_cycles}):\n> ").strip())
        if requested_cycles < 1 or requested_cycles > max_available_cycles:
            raise ValueError
    except ValueError:
        print(f"\n❌ ERROR: Please enter a valid number of cycles between 1 and {max_available_cycles}.")
        sys.exit(1)

    end_row = start_row + requested_cycles - 1
        
    print("\n" + "=" * 60)
    print(" ⚠️ CRITICAL STEP REQUIRED:")
    print(" 1. Arrange your desktop so you can see this terminal and your browser side-by-side.")
    print(" 2. Click inside the browser grid cell that matches your starting row choice")
    print(f"    (e.g., if starting at Excel Row {start_row}, click into Webpage Row {start_row}, Column 1).")
    print("=" * 60)
    
    start_input = input(f"Is your browser cursor placed inside Row {start_row}, Column 1? (yes/no): ").strip().lower()
    
    if start_input not in ['yes', 'y']:
        print("\n🛑 Execution aborted. Setup your layout windows and try running the script again.")
        sys.exit(0)
        
    print("\n⏳ Starting in 5 seconds! Immediately click back onto your web page field...")
    for countdown in range(5, 0, -1):
        print(f"   🚀 Commencing in {countdown}...")
        time.sleep(1)
        
    print(f"\n⚡ Processing action cycles on sheet '{selected_sheet}'... Keep your hands off the keyboard/mouse.")

    # Enable Desktop Emergency Fail-Safe (slam cursor into any corner to kill script execution)
    pyautogui.FAILSAFE = True

    # Outer loop: Runs the required number of cycles specified by the user (up to 70)
    for current_cycle in range(requested_cycles):
        excel_row = start_row + current_cycle
        print(f"\n🟩 [Cycle {current_cycle + 1}/{requested_cycles}] Processing Excel Row {excel_row}...")

        # --- INNER LOOP: Copies exactly 6 cells horizontally (Columns A to F) ---
        for col_idx in range(1, 7):  # Columns 1 to 6 (A to F)
            col_letter = get_column_letter(col_idx)
            
            print(f"   ➡️ Row Focus on Cell {col_letter}{excel_row}...")

            # Determine whether the current field column is targeted for an update
            if col_idx in target_columns:
                # Wipe field content (Ctrl+A -> Backspace) to ensure a clean slate before writing
                print(f"      🧹 Clearing existing field entry...")
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.04)
                pyautogui.press('backspace')
                time.sleep(0.04)

                # Fetch text string values directly out of active cells
                cell_value = sheet.cell(row=excel_row, column=col_idx).value
                text_to_paste = "" if cell_value is None else str(cell_value).strip()
                
                print(f"      📝 [UPDATE] Writing '{text_to_paste}' into Column {col_idx}")
                if text_to_paste:
                    pyautogui.write(text_to_paste, interval=0.01)
                time.sleep(0.15)  # Micro delay to allow browser UI/dropdowns to settle
                
                # Check for special Column 2 Autocomplete Selector Trigger combo
                if col_idx == 2:
                    print("      ⌨️ Dropdown cell action: Pressing DOWN ➡️ ENTER ➡️ TAB...")
                    pyautogui.press('down')
                    time.sleep(0.08)
                    pyautogui.press('enter')
                    time.sleep(0.08)
                    pyautogui.press('tab')
                    time.sleep(0.1)
                    continue  # TAB macro already handled layout movement, skip to navigation block
            else:
                # User intentionally skipped this specific field column - NO PASTE OR CLEAR ACTION PERFORMD
                print(f"      ⏭️ [SKIP] Ignoring Column {col_idx}. Executing TAB navigation action only.")
                time.sleep(0.05)

            # --- UNIFIED GRID TRANSITION NAVIGATION CONTROLS ---
            # Evaluate if the active cell is the final column item (Column F / 6)
            if col_idx == 6:
                # If this isn't the absolute final cycle, jump down to the next dashboard grid line
                if current_cycle < (requested_cycles - 1):
                    print("      ⌨️ Row end reached. Pressing TAB 4 times to bypass action buttons...")
                    for _ in range(4):
                        pyautogui.press('tab')
                        time.sleep(0.05)
            else:
                # Standard linear progression step (Columns 1, 3, 4, 5, or untargeted Column 2)
                pyautogui.press('tab')
                
            time.sleep(0.1)  # Micro-stabilization pause

        print(f"   ✓ Completed Cycle.")

    print(f"\n🎉 Success! Run completed on sheet '{selected_sheet}'. Processed {requested_cycles} rows from Row {start_row} to Row {end_row}.")
    print("=" * 60)

except Exception as error:
    print(f"\n❌ A system error occurred during automation: {error}")
