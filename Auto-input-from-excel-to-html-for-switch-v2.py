import os
import sys
import time
import openpyxl
import pyautogui
from openpyxl.utils import get_column_letter

print("=" * 60)
print("          6-COLUMN KEYBOARD MATRIX CONTROLLER")
print("=" * 60)

# 1. Dynamically request the spreadsheet file path
excel_file = input("Enter the full path to your Excel file (e.g., C:\\path\\to\\data.xlsx):\n> ").strip().strip('"\'')

if not os.path.exists(excel_file):
    print(f"\n❌ ERROR: The specified Excel file could not be found at:\n   {excel_file}")
    sys.exit(1)

try:
    # 2. Open spreadsheet using live workbook cell tracking
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # --- Display available sheets and request choice ---
    print("\n📋 Available sheets in this workbook:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {idx}. {sheet_name}")
    print("-" * 60)
    
    selected_sheet = input("Enter the exact name of the sheet you are working on:\n> ").strip()
    
    if selected_sheet not in wb.sheetnames:
        print(f"\n❌ ERROR: '{selected_sheet}' is not a valid sheet name in this workbook.")
        sys.exit(1)
        
    sheet = wb[selected_sheet]

    # 3. Dynamic row selection prompt
    try:
        start_row = int(input("\nWhich Excel row would you like to start copying from? (1-24):\n> ").strip())
        if start_row < 1 or start_row > 24:
            raise ValueError
    except ValueError:
        print("\n❌ ERROR: Please enter a valid whole number between 1 and 24.")
        sys.exit(1)
        
    # 4. Dynamic cycle requirement prompt
    max_available_cycles = 24 - start_row + 1
    try:
        requested_cycles = int(input(f"\nHow many rows/cycles do you require to run? (Max available: {max_available_cycles}):\n> ").strip())
        if requested_cycles < 1 or requested_cycles > max_available_cycles:
            raise ValueError
    except ValueError:
        print(f"\n❌ ERROR: Please enter a valid number of cycles between 1 and {max_available_cycles}.")
        sys.exit(1)

    end_row = start_row + requested_cycles - 1
        
    print("\n" + "=" * 60)
    print(" ⚠️ CRITICAL STEP REQUIRED:")
    print(" 1. Arrange your desktop so you can see this terminal and your browser side-by-side.")
    print(" 2. Click inside the browser grid cell that matches your starting row choice")
    print(f"    (e.g., if starting at Excel Row {start_row}, click into Webpage Row {start_row}, Column 1).")
    print("=" * 60)
    
    start_input = input(f"Is your browser cursor placed inside Row {start_row}, Column 1? (yes/no): ").strip().lower()
    
    if start_input not in ['yes', 'y']:
        print("\n🛑 Execution aborted. Setup your layout windows and try running the script again.")
        sys.exit(0)
        
    print("\n⏳ Starting in 5 seconds! Immediately click back onto your web page field...")
    for countdown in range(5, 0, -1):
        print(f"   🚀 Commencing in {countdown}...")
        time.sleep(1)
        
    print(f"\n⚡ Processing action cycles on sheet '{selected_sheet}'... Keep your hands off the keyboard/mouse.")

    # Outer loop: Runs the required number of cycles specified by the user
    for current_cycle in range(requested_cycles):
        excel_row = start_row + current_cycle
        print(f"\n🟩 [Cycle {current_cycle + 1}/{requested_cycles}] Processing Excel Row {excel_row}...")

        # --- INNER LOOP: Copies exactly 6 cells horizontally (Columns A to F) ---
        for col_idx in range(1, 7):  # Columns 1 to 6 (A to F)
            col_letter = get_column_letter(col_idx)
            
            # Pull live cell string data from the chosen sheet
            cell_value = sheet.cell(row=excel_row, column=col_idx).value
            text_to_paste = "" if cell_value is None else str(cell_value).strip()
            
            print(f"   ➡️ Reading Excel [{selected_sheet}] cell {col_letter}{excel_row}: '{text_to_paste}'")
            
            # 1. Paste text value directly into the web input where your cursor currently sits
            if text_to_paste:
                pyautogui.write(text_to_paste, interval=0.01)
                
            time.sleep(0.15)  # Micro delay to allow autocomplete/dropdown UI to appear
            
            # 2. Browser Focus Navigation Logic:
            # Check if this is the 2nd input box (Column B) to execute the combo dropdown sequence
            if col_idx == 2:
                print("      ⌨️ Dropdown cell detected. Pressing DOWN ➡️ ENTER ➡️ TAB...")
                pyautogui.press('down')
                time.sleep(0.08)
                pyautogui.press('enter')
                time.sleep(0.08)
                pyautogui.press('tab')
                
            # Check if we just filled out the final 6th column cell (Column F)
            elif col_idx == 6:
                # If this isn't the absolute final cycle, hop down to the next row
                if current_cycle < (requested_cycles - 1):
                    print("      ⌨️ Row end reached. Pressing TAB 4 times to bypass action buttons...")
                    for _ in range(4):
                        pyautogui.press('tab')
                        time.sleep(0.05)
            else:
                # Standard step forward between columns 1, 3, 4, and 5
                pyautogui.press('tab')
                
            time.sleep(0.1)  # Micro stabilization pause

        print(f"   ✓ Completed Cycle. Simulating Excel Enter key (moving to next Excel row)...")

    print(f"\n🎉 Success! Run completed on sheet '{selected_sheet}'. Processed {requested_cycles} rows from Row {start_row} to Row {end_row}.")
    print("=" * 60)

except Exception as error:
    print(f"\n❌ A system error occurred during automation: {error}")
